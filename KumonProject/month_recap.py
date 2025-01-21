from openpyxl import load_workbook
import pandas as pd
from datetime import datetime

FILE_NAME = "Student Logins.xlsx"
SHEET_EXPORT_NAME = "Monthly Recap"
FILE_EXTRACT_NAME = 'Student Database.xlsx'


def recap(file_name):
    students_info = {
        # Absences, avg. time in class
    }
    wb = load_workbook(file_name)
    ws = load_workbook(FILE_EXTRACT_NAME)["Tracker"]

    for student in ws:
        if student[6].value is not None and student[7].value is not None and student[6].value is not "First Name":
            students_info[student[6].value.split("(")[0] + " " + student[7].value] = {"Absences": 0, "Average Duration": []}
    for sheet in wb:
        for student in sheet:
            if student[2].value == "ABSENT":
                students_info[student[1].value]["Absences"] += 1
            else:
                try:
                    diff = (datetime.strptime(str(student[3].value), '%H:%M') -
                            datetime.strptime(str(student[2].value), '%H:%M'))
                    students_info[student[1].value]["Average Duration"].append(diff.total_seconds()//60)
                except:
                    continue

    for student in students_info:
        duration_list = students_info[student]["Average Duration"]
        if len(duration_list) > 0:
            avg = sum(duration_list) // len(duration_list)
        else:
            avg = 0.00
        students_info[student]["Average Duration"] = round(avg, 2)

    export_list = [["Name", "Absences", "Average Duration"]]

    for student in students_info:
        info = list(students_info[student].values())
        info.insert(0, student)
        export_list.append(info)

    export_list.pop(1)

    df = pd.DataFrame(export_list)

    with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
        try:
            df.to_excel(writer, sheet_name=SHEET_EXPORT_NAME, index=False)
        except ValueError:
            print("Monthly recap already exists")


if __name__ == '__main__':
    print("Are you sure you want a monthly recap?")
    answer = input("Y/N")
    if answer == "Y" or answer == "y":
        recap(FILE_NAME)
